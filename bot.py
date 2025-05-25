import os
import threading
import time
import requests
import logging
import pandas as pd
import re
import psycopg2
import tempfile
from telegram import Update, ReplyKeyboardMarkup, BotCommand, InputFile
from telegram.ext import Application, CommandHandler, MessageHandler, ContextTypes, filters
from mail_reader import start_mail_checking, ensure_database_exists
from collections import defaultdict
from datetime import datetime, timedelta

TOKEN = os.getenv("TELEGRAM_TOKEN")
ADMIN_CHAT_ID = os.getenv("ADMIN_CHAT_ID")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def get_pg_connection():
    return psycopg2.connect(
        host=os.getenv("POSTGRES_HOST"),
        port=os.getenv("POSTGRES_PORT", 5432),
        dbname=os.getenv("POSTGRES_DB"),
        user=os.getenv("POSTGRES_USER"),
        password=os.getenv("POSTGRES_PASSWORD")
    )

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    sticker_id = "CAACAgIAAxkBAAIC6mgUWmOtztmC0dnqI3C2l4wcikA-AAJvbAACa_OZSGYOhHaiIb7mNgQ"
    await update.message.reply_sticker(sticker_id)
    await update.message.reply_text("⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦"
        )
await update.message.reply_text("Привет! Отправь мне номер контейнера для отслеживания.")

async def handle_sticker(update: Update, context: ContextTypes.DEFAULT_TYPE):
    sticker = update.message.sticker
    await update.message.reply_text(
        f"🆔 ID этого стикера:\n`{sticker.file_id}`",
        parse_mode='Markdown'
    )

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_input = update.message.text
    container_numbers = [c.strip().upper() for c in re.split(r'[\s,\n.]+' , user_input.strip()) if c]

    conn = get_pg_connection()
    cursor = conn.cursor()

    found_rows = []
    not_found = []

    for number in container_numbers:
        cursor.execute("""
            SELECT container_number, from_station, to_station, current_station,
                   operation, operation_date, waybill, km_left, forecast_days,
                   wagon_number, operation_road
            FROM tracking WHERE container_number = %s
        """, (number,))
        row = cursor.fetchone()
        if row:
            found_rows.append(row)

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS stats (
                    id SERIAL PRIMARY KEY,
                    container_number TEXT,
                    user_id BIGINT,
                    username TEXT,
                    timestamp TIMESTAMP DEFAULT NOW()
                )
            """)
            cursor.execute("""
                INSERT INTO stats (container_number, user_id, username)
                VALUES (%s, %s, %s)
            """, (number, update.message.from_user.id, update.message.from_user.username))
            conn.commit()
        else:
            not_found.append(number)

    conn.close()

    if len(container_numbers) > 1 and found_rows:
        df = pd.DataFrame(found_rows, columns=[
            'Номер контейнера', 'Станция отправления', 'Станция назначения',
            'Станция операции', 'Операция', 'Дата и время операции',
            'Номер накладной', 'Расстояние оставшееся', 'Прогноз прибытия (дней)',
            'Номер вагона', 'Дорога операции'
        

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Дислокация')
                workbook = writer.book
                worksheet = writer.sheets['Дислокация']
            
                # Заливка для шапки
                from openpyxl.styles import PatternFill
                fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
                for cell in worksheet[1]:
                    cell.fill = fill
            
                # Автоширина столбцов
                for col in worksheet.columns:
                    max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                    adjusted_width = max_length + 2
                    worksheet.column_dimensions[col[0].column_letter].width = adjusted_width

            from datetime import datetime, timedelta

            vladivostok_time = datetime.utcnow() + timedelta(hours=10)
            filename = f"Дислокация {vladivostok_time.strftime('%H-%M')}.xlsx"
            await update.message.reply_document(document=open(tmp.name, "rb"), filename=filename)

        if not_found:
            await update.message.reply_text("⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦"
        )
"❌ Не найдены: " + ", ".join(not_found))
        return

    if found_rows:
        reply_lines = []
        for row in found_rows:
            wagon_type = "полувагон" if row[9].startswith("6") else "платформа"
            reply_lines.append(
                f"🚛 Контейнер: {row[0]}\n"
                f"🚇 Вагон: {row[9]} {wagon_type}\n"
                f"📍Дислокация: {row[3]} {row[10]}\n"
                f"🏗 Операция: {row[4]}\n📅 {row[5]}\n\n"
                f"Откуда: {row[1]}\nКуда: {row[2]}\n\n"
                f"Накладная: {row[6]}\nОсталось км: {row[7]}\n"
                f"📅 Прогноз прибытия: {row[8]} дн."
            )
        await update.message.reply_text("⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦"
        )
"\n" + "═" * 30 + "\n".join(reply_lines))
    else:
        await update.message.reply_text("⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦"
        )
"Ничего не найдено по введённым номерам.")

async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if str(update.message.chat_id) != ADMIN_CHAT_ID:
        await update.message.reply_text("⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦"
        )
"⛔ Доступ запрещён.")
        return

    conn = get_pg_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT user_id, COALESCE(username, '—') AS username, COUNT(*) AS запросов,
               STRING_AGG(DISTINCT container_number, ', ') AS контейнеры
        FROM stats
        WHERE timestamp >= NOW() - INTERVAL '1 day'
          AND user_id != 114419850
        GROUP BY user_id, username
        ORDER BY запросов DESC
    """)
    rows = cursor.fetchall()
    conn.close()

    if not rows:
        await update.message.reply_text("⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦"
        )
"Нет статистики за последние сутки.")
        return

    text = "📊 Статистика за последние 24 часа:\n\n"
    messages = []
    for row in rows:
        entry = (
            f"👤 {row[1]} (ID: {row[0]})\n"
            f"Запросов: {row[2]}\n"
            f"Контейнеры: {row[3]}\n\n"
        )
        if len(text) + len(entry) > 4000:
            messages.append(text)
            text = ""
        text += entry
    messages.append(text)

    for msg in messages:
        await update.message.reply_text("⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦"
        )
msg)

async def exportstats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if str(update.message.chat_id) != ADMIN_CHAT_ID:
        await update.message.reply_text("⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦"
        )
"⛔ Доступ запрещён.")
        return

    conn = get_pg_connection()
    df = pd.read_sql_query("SELECT * FROM stats", conn)
    conn.close()

    if df.empty:
        await update.message.reply_text("⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦"
        )
"Нет данных для экспорта.")
        return

    from openpyxl.styles import PatternFill
    from datetime import datetime, timedelta
    import tempfile

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Статистика')
            workbook = writer.book
            worksheet = writer.sheets['Статистика']

            # Заливка шапки таблицы
            header_fill = PatternFill(start_color='FFD673', end_color='FFD673', fill_type='solid')
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Автоширина столбцов
            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                worksheet.column_dimensions[col[0].column_letter].width = max_length + 2

        # Имя файла с учетом Владивостокского времени
        vladivostok_time = datetime.utcnow() + timedelta(hours=10)
        filename = f"Статистика {vladivostok_time.strftime('%H-%M')}.xlsx"
        await update.message.reply_document(document=open(tmp.name, "rb"), filename=filename)

    
        BotCommand("start", "Начать работу с ботом"),
        BotCommand("stats", "Статистика запросов (для администратора)"),
        BotCommand("exportstats", "Выгрузка всех запросов в Excel (админ)")
    

def ensure_database_exists():
    conn = get_pg_connection()
    cursor = conn.cursor()

    # Создание таблицы tracking
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS tracking (
            container_number TEXT,
            from_station TEXT,
            to_station TEXT,
            current_station TEXT,
            operation TEXT,
            operation_date TEXT,
            waybill TEXT,
            km_left TEXT,
            forecast_days TEXT,
            wagon_number TEXT,
            operation_road TEXT
        );
    """)

    # Создание таблицы stats
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS stats (
            id SERIAL PRIMARY KEY,
            container_number TEXT,
            user_id BIGINT,
            username TEXT,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)

    conn.commit()
    conn.close()

def keep_alive():
    """Периодически пингует Render, чтобы не засыпал."""
    url = f"https://{os.environ.get('RENDER_EXTERNAL_HOSTNAME')}/"
    def ping():
        while True:
            try:
                response = requests.get(url)
                logger.info(f"[AUTOPING] {url} — {response.status_code}")
            except Exception as e:
                logger.warning(f"[AUTOPING] Error: {e}")
            time.sleep(600)  # каждые 10 минут
    threading.Thread(target=ping, daemon=True).start()


async def broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if str(update.message.chat_id) != ADMIN_CHAT_ID:
        await update.message.reply_text("⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦"
        )
"⛔ Доступ запрещён.")
        return

    if not context.args:
        await update.message.reply_text("⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦"
        )
"⚠️ Укажи текст:\n/broadcast Обновление в боте 📦")
        return

    message = " ".join(context.args)

    conn = get_pg_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT user_id FROM stats WHERE user_id IS NOT NULL")
    user_ids = [row[0] for row in cursor.fetchall()]
    conn.close()

    success, failed = 0, 0
    for uid in user_ids:
        try:
            await context.bot.send_message(chat_id=uid, text=message)
            success += 1
        except Exception as e:
            logger.warning(f"❌ Не отправлено {uid}: {e}")
            failed += 1

    await update.message.reply_text("⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦"
        )
f"📤 Рассылка завершена:\n✅ {success} успешно\n❌ {failed} с ошибкой")


from telegram import BotCommandScopeDefault, BotCommandScopeChat

async def set_bot_commands(application):
    # Команды для всех пользователей
    public_commands = [
        BotCommand("start", "Начать работу с ботом"),
    ]
    await application.bot.set_my_commands(public_commands, scope=BotCommandScopeDefault())

    # Команды только для админа
    admin_commands = [
        BotCommand("stats", "Статистика запросов"),
        BotCommand("exportstats", "Выгрузка всех запросов в Excel"),
        BotCommand("broadcast", "Предпросмотр рассылки"),
        BotCommand("broadcast_confirm", "Подтверждение и отправка")
    ]
    await application.bot.set_my_commands(admin_commands, scope=BotCommandScopeChat(chat_id=int(ADMIN_CHAT_ID)))



async def broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if str(update.message.chat_id) != ADMIN_CHAT_ID:
        await update.message.reply_text("⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦"
        )
"⛔ Доступ запрещён.")
        return

    if not context.args:
        await update.message.reply_text("⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦"
        )

            "⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦
        )
        return

    message = " ".join(context.args)

    # Сначала отправим сообщение только админу для подтверждения
    preview_text = f"🔍 Предпросмотр рассылки:

{message}

Если всё ок — отправь /broadcast_confirm"
    await context.bot.send_message(chat_id=ADMIN_CHAT_ID, text=preview_text)

    # Сохраняем сообщение во временное хранилище (in-memory)
    context.bot_data["broadcast_pending"] = message


async def broadcast_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if str(update.message.chat_id) != ADMIN_CHAT_ID:
        await update.message.reply_text("⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦"
        )
"⛔ Доступ запрещён.")
        return

    message = context.bot_data.get("broadcast_pending")
    if not message:
        await update.message.reply_text("⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦"
        )
"❌ Нет сообщения для подтверждённой рассылки.")
        return

    conn = get_pg_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT user_id FROM stats WHERE user_id IS NOT NULL")
    user_ids = [row[0] for row in cursor.fetchall()]
    conn.close()

    success, failed = 0, 0
    for uid in user_ids:
        try:
            await context.bot.send_message(chat_id=uid, text=message)
            success += 1
        except Exception as e:
            logger.warning(f"❌ Не отправлено {uid}: {e}")
            failed += 1

    await update.message.reply_text("⚠️ Укажи текст для отправки:\n\n/broadcast Это тестовое сообщение 📦"
        )
f"📤 Рассылка завершена:\n✅ {success} успешно\n❌ {failed} с ошибкой")
    context.bot_data.pop("broadcast_pending", None)


def main():
    ensure_database_exists()
    start_mail_checking()

    keep_alive()
    application = Application.builder().token(TOKEN).build()
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("stats", stats))
    application.add_handler(CommandHandler("exportstats", exportstats))
    application.add_handler(CommandHandler("broadcast", broadcast))
    application.add_handler(CommandHandler("broadcast_confirm", broadcast_confirm))
    application.add_handler(MessageHandler(filters.Sticker.ALL, handle_sticker))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    application.post_init = set_bot_commands
    logger.info("✨ Бот запущен!")
    application.run_webhook(
        listen="0.0.0.0",
        port=int(os.environ.get("PORT", 10000)),
        url_path=TOKEN,
        webhook_url=f"https://{os.environ.get('RENDER_EXTERNAL_HOSTNAME')}/{TOKEN}"
    )

if __name__ == '__main__':
    main()

