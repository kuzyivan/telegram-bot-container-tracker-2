import pandas as pd
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import tempfile
import re

def create_excel_file(rows, columns):
    """
    Однолистовой Excel-файл.
    rows: список списков с данными
    columns: список названий столбцов
    """
    df = pd.DataFrame(rows, columns=columns)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Экспорт')
            worksheet = writer.sheets['Экспорт']
            header_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
            for cell in worksheet[1]:
                cell.fill = header_fill
            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                worksheet.column_dimensions[col[0].column_letter].width = max_length + 2
        return tmp.name

def clean_sheet_name(name):
    return re.sub(r'[:\\/?*\[\]]', '_', str(name))[:31]

def create_excel_multisheet(data_per_user, columns):
    import pandas as pd
    from openpyxl.styles import PatternFill
    import tempfile

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
            for user_label, rows in data_per_user.items():
                sheet_name = clean_sheet_name(user_label)
                df = pd.DataFrame(rows, columns=columns)
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                worksheet = writer.sheets[sheet_name]
                header_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
                for cell in worksheet[1]:
                    cell.fill = header_fill
                for col in worksheet.columns:
                    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                    worksheet.column_dimensions[col[0].column_letter].width = max_length + 2
        return tmp.name

def get_vladivostok_filename(prefix="Слежение контейнеров"):
    """
    Генерирует имя файла по Владивостокскому времени (UTC+10).
    """
    vladivostok_time = datetime.utcnow() + timedelta(hours=10)
    return f"{prefix} {vladivostok_time.strftime('%H-%M')}.xlsx"
