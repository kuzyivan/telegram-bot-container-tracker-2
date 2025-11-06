# utils/send_tracking.py
import os
import pandas as pd
from io import BytesIO
from datetime import datetime
from typing import List, Any
import logging
from pytz import timezone 
import uuid 
import openpyxl 
from pandas import ExcelWriter 
import openpyxl.utils
# <--- НОВЫЙ ИМПОРТ для стилей
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

def create_excel_file(rows: List[List[Any]], columns: List[str]) -> str:
    """Создает временный Excel-файл из списка строк и заголовков с форматированием."""
    logger.info(f"Создание Excel-файла (один лист) с форматированием, строк: {len(rows)}")

    # Создаем DataFrame
    df = pd.DataFrame(rows, columns=columns)
    
    # --- ✅ НОВОЕ: Находим колонки, которые содержат даты ---
    # Мы будем искать их по имени, т.к. они уже могут быть datetime объектами
    date_columns_names = [
        'Дата отправления', 
        'Дата и время операции', 
        'Начало рейса' # Добавим на всякий случай
    ]
    
    # Преобразуем колонки в datetime (если они еще не) и убираем таймзону (Excel ее не любит)
    for col_name in date_columns_names:
        if col_name in df.columns:
            try:
                df[col_name] = pd.to_datetime(df[col_name]).dt.tz_localize(None)
            except Exception as e:
                logger.warning(f"Не удалось преобразовать колонку '{col_name}' в datetime: {e}")
                
    # --- КОНЕЦ НОВОГО БЛОКА ---

    # Создаем временный файл
    temp_file_path = os.path.join('/tmp', f'tmp{uuid.uuid4().hex}.xlsx')
    
    try:
        # Используем pandas.ExcelWriter
        with ExcelWriter(temp_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Дислокация')
            
            # --- ЛОГИКА ФОРМАТИРОВАНИЯ ---
            workbook = writer.book
            worksheet = writer.sheets['Дислокация']
            
            # 1. Форматирование заливки заголовка
            fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            
            # 2. Форматирование выравнивания и жирности
            header_font = Font(bold=True)
            
            # 3. Применение форматирования к первой строке (заголовкам)
            for cell in worksheet["1:1"]:
                cell.fill = fill
                cell.font = header_font
                
            # --- ✅ НОВОЕ: Применяем формат даты к колонкам ---
            # Задаем формат ДД.ММ.ГГГГ ЧЧ:ММ
            date_format_style = 'DD.MM.YYYY HH:MM' 
            
            # 4. Установка ширины столбцов и формата дат
            for col_idx, column_name in enumerate(columns, 1):
                max_length = len(column_name)
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                
                # Логика подбора ширины
                adjusted_width = max_length * 1.5 + 5
                
                # Применяем формат даты, если колонка найдена
                if column_name in date_columns_names:
                    logger.debug(f"Применяю формат даты к колонке {column_letter} ({column_name})")
                    for cell in worksheet[column_letter]:
                        if cell.row > 1: # Пропускаем заголовок
                            cell.number_format = date_format_style
                    adjusted_width = 20 # Фиксированная ширина для дат
                
                worksheet.column_dimensions[column_letter].width = max(15, min(adjusted_width, 50))
            
            # --- КОНЕЦ НОВОГО БЛОКА ---
            
        logger.info(f"Excel-файл успешно создан и отформатирован: {temp_file_path}")
        return temp_file_path
    
    except Exception as e:
        logger.error(f"❌ Ошибка при создании Excel-файла: {e}", exc_info=True)
        raise ValueError(e)

def get_vladivostok_filename(prefix: str) -> str:
    """Возвращает имя файла с текущей датой и временем по Владивостоку."""
    tz = timezone('Asia/Vladivostok')
    now = datetime.now(tz)
    return f"{prefix}_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"